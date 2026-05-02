"""
streamlit_app.py
----------------
Full-featured web UI for Smart AI File Organizer.

Features
--------
  - Upload any supported file; AI classifies instantly
  - Confidence scores with colour coding
  - Language detection
  - Content preview
  - Manual override (correct category)
  - AI Smart Rename (NVIDIA API)
  - Category Manager (add/edit/delete categories)
  - Export results as Excel
  - Try with your own text
"""

import csv
import io
import json
import os
import tempfile
from pathlib import Path

import streamlit as st

ROOT = Path(__file__).resolve().parent.parent

# -- Page config ---------------------------------------------------------------
st.set_page_config(
    page_title="Smart AI File Organizer",
    page_icon="SAFO",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -- CSS -----------------------------------------------------------------------
st.markdown("""
<style>
.category-badge {
    display: inline-block;
    padding: 8px 24px;
    border-radius: 20px;
    font-size: 1.2rem;
    font-weight: 700;
    margin: 8px 0;
}
.preview-box {
    background: #1e1e2e;
    border: 1px solid #2a2a3e;
    border-radius: 10px;
    padding: 14px;
    font-size: 0.85rem;
    color: #94a3b8;
    font-family: monospace;
    white-space: pre-wrap;
    max-height: 200px;
    overflow-y: auto;
}
.result-card {
    background: #1e1e2e;
    border: 1px solid #2a2a3e;
    border-radius: 12px;
    padding: 20px;
    margin: 10px 0;
}
</style>
""", unsafe_allow_html=True)

# -- Constants -----------------------------------------------------------------
CAT_COLORS = {
    "Finance": "#ef4444", "Resume": "#3b82f6", "AI": "#10b981",
    "Research": "#f97316", "Personal": "#8b5cf6", "Legal": "#ec4899",
    "Medical": "#06b6d4", "Other": "#64748b",
}
CAT_ICONS = {
    "Finance": "[FIN]", "Resume": "[CV]", "AI": "[AI]", "Research": "[RES]",
    "Personal": "[PER]", "Legal": "[LAW]", "Medical": "[MED]", "Other": "[GEN]",
}
SUPPORTED = [".pdf", ".txt", ".docx", ".xlsx", ".pptx",
             ".csv", ".eml", ".msg", ".zip", ".png", ".jpg", ".jpeg"]
CONFIG_PATH = ROOT / "config.example.json"

# -- Session state init --------------------------------------------------------
if "results" not in st.session_state:
    st.session_state.results = []   # list of dicts
if "clf" not in st.session_state:
    st.session_state.clf = None

# -- Load classifier -----------------------------------------------------------
@st.cache_resource(show_spinner="Loading AI model...")
def load_classifier():
    from .classifier import DocumentClassifier
    clf = DocumentClassifier()
    clf.train()
    return clf

@st.cache_resource(show_spinner=False)
def get_lang_tools():
    from .classifier import LANGUAGE_NAMES, detect_language
    return detect_language, LANGUAGE_NAMES

def get_clf():
    if st.session_state.clf is None:
        st.session_state.clf = load_classifier()
    return st.session_state.clf

# -- Sidebar -------------------------------------------------------------------
with st.sidebar:
    st.markdown("## Smart AI File Organizer")
    st.markdown("AI-powered document classification with confidence scores, manual override, and more.")
    st.divider()

    page = st.radio("Navigate", [
        "Classify Files",
        "Try with Text",
        "Results & Export",
        "Semantic Search",
        "Category Manager",
    ], label_visibility="collapsed")

    st.divider()

    # AI Smart Rename toggle
    st.markdown("### AI Smart Rename")
    smart_rename = st.toggle("Enable Smart Rename", value=False)
    nvidia_key = ""
    if smart_rename:
        nvidia_key = st.text_input("NVIDIA API Key", type="password",
                                    placeholder="nvapi-...",
                                    help="Get free key at build.nvidia.com/models")

    st.divider()
    st.markdown("[![GitHub](https://img.shields.io/badge/GitHub-Source-black?logo=github)](https://github.com/sarawagh27/smart-ai-file-organizer)")

# ------------------------------------------------------------------------------
# PAGE 1 - Classify Files
# ------------------------------------------------------------------------------
if page == "Classify Files":
    st.markdown("## Classify Files")
    st.markdown("Upload one or more files - AI reads the content and classifies each one.")

    uploaded_files = st.file_uploader(
        "Drop files here",
        type=[s.lstrip(".") for s in SUPPORTED],
        accept_multiple_files=True,
        help=f"Supported: {', '.join(SUPPORTED)}",
    )

    if uploaded_files:
        if st.button("Classify All", type="primary", use_container_width=True):
            clf = get_clf()
            detect_language, LANGUAGE_NAMES = get_lang_tools()

            from .renamer import SmartRenamer
            from .text_extractor import extract_text
            renamer = SmartRenamer(api_key=nvidia_key, enabled=smart_rename and bool(nvidia_key))

            progress = st.progress(0, text="Classifying...")
            new_results = []

            for i, uploaded in enumerate(uploaded_files):
                ext = Path(uploaded.name).suffix.lower()
                progress.progress((i + 1) / len(uploaded_files),
                                  text=f"Processing {uploaded.name}...")

                with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                    tmp.write(uploaded.read())
                    tmp_path = tmp.name

                try:
                    text = extract_text(tmp_path)
                    os.unlink(tmp_path)

                    category, confidence, is_low = clf.predict_with_confidence(text)
                    lang_code = detect_language(text)
                    lang_name = LANGUAGE_NAMES.get(lang_code, lang_code.upper())

                    # Smart rename
                    new_name = renamer.rename(uploaded.name, text) if smart_rename else uploaded.name

                    new_results.append({
                        "filename":     uploaded.name,
                        "new_name":     new_name,
                        "category":     category,
                        "confidence":   confidence,
                        "is_low":       is_low,
                        "language":     lang_name,
                        "words":        len(text.split()),
                        "preview":      " ".join(text.split())[:300] + ("..." if len(text.split()) > 50 else ""),
                        "overridden":   False,
                    })
                except Exception as e:
                    new_results.append({
                        "filename": uploaded.name, "new_name": uploaded.name,
                        "category": "Error", "confidence": 0.0,
                        "is_low": True, "language": "Unknown",
                        "words": 0, "preview": str(e), "overridden": False,
                    })

            progress.empty()
            st.session_state.results.extend(new_results)
            st.success(f"Classified {len(new_results)} file(s)!")
            st.info("Go to **Results & Export** to see results, override categories, and export.")

# ------------------------------------------------------------------------------
# PAGE 2 - Try with Text
# ------------------------------------------------------------------------------
elif page == "Try with Text":
    st.markdown("## Try with Text")
    st.markdown("Paste any text and see how the AI classifies it.")

    text_input = st.text_area("Paste text:", height=180,
                               placeholder="Paste document text here...")

    if st.button("Classify", type="primary", use_container_width=True):
        if not text_input.strip() or len(text_input.strip()) < 20:
            st.warning("Please enter at least 20 characters.")
        else:
            with st.spinner("Classifying..."):
                clf = get_clf()
                detect_language, LANGUAGE_NAMES = get_lang_tools()
                category, confidence, is_low = clf.predict_with_confidence(text_input)
                lang_code = detect_language(text_input)
                lang_name = LANGUAGE_NAMES.get(lang_code, lang_code.upper())

            color = CAT_COLORS.get(category, "#64748b")
            icon  = CAT_ICONS.get(category, "[GEN]")

            col1, col2, col3 = st.columns(3)
            col1.metric("Category", f"{icon} {category}")
            col2.metric("Confidence", f"{confidence:.1f}%")
            col3.metric("Language", lang_name)

            st.progress(int(min(confidence, 100)),
                        text=f"{'Low confidence' if is_low else 'Confident'}")

            if is_low:
                st.warning("Low confidence - try adding more descriptive text.")

            # Manual override for text
            st.divider()
            st.markdown("**Not right? Override the category:**")
            cats = list(CAT_COLORS.keys())
            override = st.selectbox("Correct category:", ["- keep as is -"] + cats)
            if override != "- keep as is -" and st.button("Apply Override"):
                clf.add_correction(text_input, override)
                st.success(f"Correction saved! Model will now learn '{override}' for this text.")

# ------------------------------------------------------------------------------
# PAGE 3 - Results & Export
# ------------------------------------------------------------------------------
elif page == "Results & Export":
    st.markdown("## Results")

    if not st.session_state.results:
        st.info("No results yet. Go to **Classify Files** to classify some files first.")
    else:
        # Summary metrics
        results = st.session_state.results
        total     = len(results)
        low_conf  = sum(1 for r in results if r["is_low"])
        avg_conf  = sum(r["confidence"] for r in results) / total
        renamed   = sum(1 for r in results if r["new_name"] != r["filename"])

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Files", total)
        c2.metric("Avg Confidence", f"{avg_conf:.1f}%")
        c3.metric("Low Confidence", low_conf)
        c4.metric("AI Renamed", renamed)

        st.divider()

        # Results table with override
        clf = get_clf()
        cats = list(CAT_COLORS.keys())

        for i, r in enumerate(results):
            color  = CAT_COLORS.get(r["category"], "#64748b")
            icon   = CAT_ICONS.get(r["category"], "[GEN]")
            status = "Low" if r["is_low"] else "Confident"
            renamed_tag = f" -> `{r['new_name']}`" if r["new_name"] != r["filename"] else ""

            with st.expander(f"{icon} **{r['filename']}**{renamed_tag} - {r['category']} ({r['confidence']:.1f}%) {status}"):
                col1, col2 = st.columns([2, 1])

                with col1:
                    st.markdown(f"**Category:** {icon} {r['category']}")
                    st.progress(int(min(r["confidence"], 100)),
                                text=f"{r['confidence']:.1f}% confidence")
                    st.markdown(f"**Language:** {r['language']} | **Words:** {r['words']:,}")

                    if r["preview"]:
                        st.markdown("**Content Preview:**")
                        st.markdown(f'<div class="preview-box">{r["preview"]}</div>',
                                    unsafe_allow_html=True)

                with col2:
                    st.markdown("**Override Category:**")
                    new_cat = st.selectbox("", ["- keep -"] + cats,
                                           key=f"override_{i}",
                                           label_visibility="collapsed")
                    if new_cat != "- keep -" and st.button("Apply", key=f"apply_{i}"):
                        old_cat = r["category"]
                        st.session_state.results[i]["category"] = new_cat
                        st.session_state.results[i]["overridden"] = True
                        clf.add_correction(r["preview"], new_cat)
                        st.success(f"{old_cat} -> {new_cat} (model learned!)")
                        st.rerun()

        st.divider()

        # Export buttons
        col1, col2 = st.columns(2)

        with col1:
            # Export as CSV
            csv_buf = io.StringIO()
            writer  = csv.writer(csv_buf)
            writer.writerow(["File", "New Name", "Category", "Confidence (%)", "Language", "Status"])
            for r in results:
                writer.writerow([
                    r["filename"], r["new_name"], r["category"],
                    f"{r['confidence']:.1f}", r["language"],
                    "Low confidence" if r["is_low"] else "Confident",
                ])
            st.download_button("Export CSV", csv_buf.getvalue(),
                               "organizer_results.csv", "text/csv",
                               use_container_width=True)

        with col2:
            # Export as Excel
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb  = Workbook()
                ws  = wb.active
                ws.title = "Results"

                hdr_fill = PatternFill("solid", start_color="2a2a3e")
                hdr_font = Font(bold=True, color="e2e8f0", name="Segoe UI", size=10)
                thin     = Side(style="thin", color="3a3a5e")
                border   = Border(left=thin, right=thin, top=thin, bottom=thin)

                headers   = ["#", "File", "New Name", "Category", "Confidence (%)", "Language", "Status"]
                col_widths= [5, 35, 35, 14, 16, 12, 16]

                for col, (h, w) in enumerate(zip(headers, col_widths), 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                    ws.column_dimensions[get_column_letter(col)].width = w

                cat_colors_xl = {
                    "Finance":"dcfce7","Resume":"dbeafe","AI":"f0fdf4",
                    "Research":"fef9c3","Personal":"fae8ff","Legal":"fee2e2",
                    "Medical":"cffafe","Other":"f1f5f9",
                }

                for i, r in enumerate(results, 1):
                    row    = i + 1
                    bg     = cat_colors_xl.get(r["category"], "ffffff")
                    fill   = PatternFill("solid", start_color=bg)
                    rfont  = Font(name="Segoe UI", size=9)
                    status = "Low confidence" if r["is_low"] else "Confident"
                    vals   = [i, r["filename"], r["new_name"], r["category"],
                              round(r["confidence"], 1), r["language"], status]

                    for col, val in enumerate(vals, 1):
                        cell = ws.cell(row=row, column=col, value=val)
                        cell.fill   = fill
                        cell.font   = rfont
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center" if col != 2 else "left")

                ws.freeze_panes = "A2"

                xl_buf = io.BytesIO()
                wb.save(xl_buf)
                xl_buf.seek(0)

                st.download_button("Export Excel", xl_buf.getvalue(),
                                   "organizer_results.xlsx",
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
            except Exception as e:
                st.error(f"Excel export error: {e}")

        # Clear results
        st.divider()
        if st.button("Clear All Results", use_container_width=True):
            st.session_state.results = []
            st.rerun()

# ------------------------------------------------------------------------------
# PAGE 4 - Semantic Search
# ------------------------------------------------------------------------------
elif page == "Semantic Search":
    st.markdown("## Semantic Search")
    st.markdown("Search your organised files by **meaning**, not just keywords.")

    folder_path = st.text_input(
        "Path to your organised folder:",
        placeholder="e.g. D:/Downloads  or  /Users/sara/Downloads",
        help="The folder where your files have been organised into category sub-folders",
    )

    col1, col2 = st.columns([2, 1])
    with col1:
        query = st.text_input("Search query:", placeholder="e.g. medical reports blood test, invoices from Amazon, AI research papers")
    with col2:
        cat_filter = st.selectbox("Filter by category:",
                                   ["All"] + list(CAT_COLORS.keys()))

    col_a, col_b = st.columns(2)
    search_btn  = col_a.button("Search", type="primary", use_container_width=True)
    rebuild_btn = col_b.button("Rebuild Index", use_container_width=True)

    if rebuild_btn:
        if not folder_path:
            st.warning("Please enter a folder path first.")
        else:
            with st.spinner("Building search index... this may take a moment."):
                try:
                    from .search import SemanticSearch
                    engine = SemanticSearch(target_dir=folder_path)
                    n = engine.build_index(force=True)
                    st.session_state["search_engine"] = engine
                    st.success(f"Index built - {n} file(s) indexed and ready to search!")
                except Exception as e:
                    st.error(f"Error: {e}")

    if search_btn:
        if not query.strip():
            st.warning("Please enter a search query.")
        elif not folder_path:
            st.warning("Please enter a folder path.")
        else:
            # Load or build engine
            if "search_engine" not in st.session_state:
                with st.spinner("Loading index..."):
                    try:
                        from .search import SemanticSearch
                        engine = SemanticSearch(target_dir=folder_path)
                        n = engine.build_index()
                        st.session_state["search_engine"] = engine
                        if n == 0:
                            st.warning("No files indexed. Run the organiser first then rebuild.")
                            st.stop()
                    except Exception as e:
                        st.error(f"Error: {e}")
                        st.stop()

            engine = st.session_state["search_engine"]
            cf = None if cat_filter == "All" else cat_filter

            with st.spinner("Searching..."):
                results = engine.search(query, top_k=20, category_filter=cf)

            if not results:
                st.warning(f'No results for "{query}" - try different keywords or rebuild the index.')
            else:
                st.success("Found " + str(len(results)) + " result(s) for: " + query)
                st.divider()

                for fname, category, score, preview, path in results:
                    color = CAT_COLORS.get(category, "#64748b")
                    icon  = CAT_ICONS.get(category, "[GEN]")
                    bar_color = "High" if score >= 60 else "Medium" if score >= 40 else "Low"

                    with st.container():
                        c1, c2, c3 = st.columns([3, 1, 1])
                        c1.markdown(f"**{fname}**")
                        c2.markdown(f"{icon} {category}")
                        c3.markdown(f"{bar_color} {score:.1f}%")
                        st.caption(f"{preview[:120]}...")
                        st.divider()

    st.info("**How it works:** The AI reads the actual content of your files and finds ones that match your query by meaning, not just keyword matching. Works best on text-heavy files like PDFs, DOCX, and TXT.")

# ------------------------------------------------------------------------------
# PAGE 5 - Category Manager
# ------------------------------------------------------------------------------
elif page == "Category Manager":
    st.markdown("## Category Manager")
    st.markdown("Add, edit, or delete categories. Changes apply to the current session.")

    # Load config
    try:
        with open(CONFIG_PATH) as f:
            cfg = json.load(f)
    except Exception:
        cfg = {"categories": list(CAT_COLORS.keys()), "training_data": {}}

    if "session_cfg" not in st.session_state:
        st.session_state.session_cfg = cfg

    cfg = st.session_state.session_cfg
    categories   = cfg.get("categories", [])
    training_data = cfg.get("training_data", {})

    # -- View / Edit existing --------------------------------------------------
    st.markdown("### Existing Categories")

    for cat in categories:
        with st.expander(f"{CAT_ICONS.get(cat, '[GEN]')} **{cat}**"):
            keywords = training_data.get(cat, [])
            new_kw = st.text_area(
                "Keywords (one sentence per line):",
                value="\n".join(keywords),
                height=120,
                key=f"kw_{cat}",
            )
            col1, col2 = st.columns([1, 1])
            with col1:
                if st.button(f"Save", key=f"save_{cat}"):
                    lines = [l.strip() for l in new_kw.split("\n") if l.strip()]
                    st.session_state.session_cfg["training_data"][cat] = lines
                    st.success(f"Saved {len(lines)} keyword(s) for {cat}")
            with col2:
                if cat != "Other":
                    if st.button(f"Delete", key=f"del_{cat}", type="secondary"):
                        st.session_state.session_cfg["categories"].remove(cat)
                        st.session_state.session_cfg["training_data"].pop(cat, None)
                        st.rerun()

    st.divider()

    # -- Add new category ------------------------------------------------------
    st.markdown("### Add New Category")
    col1, col2 = st.columns([1, 2])
    with col1:
        new_cat_name = st.text_input("Category name:", placeholder="e.g. Engineering")
    with col2:
        new_cat_kw = st.text_area("Keywords:", placeholder="circuit board hardware firmware embedded systems", height=80)

    if st.button("Add Category", type="primary"):
        if not new_cat_name.strip():
            st.warning("Please enter a category name.")
        elif new_cat_name in categories:
            st.warning(f"'{new_cat_name}' already exists.")
        else:
            st.session_state.session_cfg["categories"].append(new_cat_name)
            kws = [l.strip() for l in new_cat_kw.split("\n") if l.strip()]
            st.session_state.session_cfg["training_data"][new_cat_name] = kws or [f"keywords for {new_cat_name}"]
            st.success(f"Added '{new_cat_name}'! Restart classifier to apply.")
            st.rerun()

    st.divider()
    st.info("After editing categories, re-classify your files to use the updated model.")
