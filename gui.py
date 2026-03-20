"""
gui.py
------
Tkinter GUI for the Smart AI File Organizer.

Level 1 additions:
  1. Confidence scores — shown in results table with colour coding
  2. Manual Override — right-click any result to change its category
  3. Dark/Light theme toggle — saves preference to config.json
"""

import csv
import json
import logging
import os
import queue
import sys
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, font, messagebox, scrolledtext, ttk

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

from organizer import FileOrganizer
from category_manager import CategoryManager
from undo import undo_moves
from utils import scan_directory
from watcher import FolderWatcher

CONFIG_PATH = ROOT / "config.json"

# ── theme palettes ──────────────────────────────────────────────────────────
THEMES = {
    "dark": {
        "BG":       "#1e1e2e",
        "SURFACE":  "#2a2a3e",
        "SURFACE2": "#313147",
        "TEXT":     "#e2e8f0",
        "MUTED":    "#64748b",
        "WHITE":    "#ffffff",
    },
}
ACCENT      = "#7c6af7"
ACCENT_DARK = "#5a48d4"
WATCH_CLR   = "#06b6d4"
WATCH_DARK  = "#0891b2"
UNDO_CLR    = "#f97316"
UNDO_DARK   = "#ea6c00"
SUCCESS     = "#4ade80"
WARNING     = "#fbbf24"
ERROR       = "#f87171"
INFO        = "#93c5fd"


class QueueHandler(logging.Handler):
    def __init__(self, log_queue):
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record):
        self.log_queue.put(record)


class SmartOrganizerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Smart AI File Organizer")
        self.geometry("920x780")
        self.minsize(760, 650)

        self._theme_name = "dark"
        self._theme      = THEMES[self._theme_name]

        self._folder    = tk.StringVar()
        self._dry_run   = tk.BooleanVar(value=True)
        self._recursive = tk.BooleanVar(value=False)
        self._smart_rename = tk.BooleanVar(value=False)
        self._running   = False
        self._watching  = False
        self._watcher   = None
        self._organizer = None   # current FileOrganizer instance
        self._log_queue : queue.Queue = queue.Queue()

        self.configure(bg=self._theme["BG"])

        self._build_fonts()
        self._build_header()
        self._build_folder_row()
        self._build_options_row()
        self._build_file_types_row()
        self._build_action_buttons()
        self._build_notebook()   # tabbed area: Log + Results
        self._build_stats_bar()
        self._setup_logging()
        self._poll_log_queue()
        self._bind_shortcuts()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── theme helpers ────────────────────────────────────────────────────────
    def _load_theme_pref(self) -> str:
        try:
            with open(CONFIG_PATH) as f:
                cfg = json.load(f)
            return cfg.get("gui", {}).get("theme", "dark")
        except Exception:
            return "dark"

    def _save_theme_pref(self, theme_name: str):
        try:
            with open(CONFIG_PATH) as f:
                cfg = json.load(f)
            cfg.setdefault("gui", {})["theme"] = theme_name
            with open(CONFIG_PATH, "w") as f:
                json.dump(cfg, f, indent=2)
        except Exception:
            pass

    def _toggle_theme(self):
        self._theme_name = "light" if self._theme_name == "dark" else "dark"
        self._theme      = THEMES[self._theme_name]
        self._save_theme_pref(self._theme_name)
        self._apply_theme()
        icon = "☀️" if self._theme_name == "light" else "🌙"
        self._btn_theme.configure(text=icon)

    def _apply_theme(self):
        t = self._theme
        self.configure(bg=t["BG"])
        # Recursively update all widgets
        self._update_widget_colors(self, t)

    def _update_widget_colors(self, widget, t):
        try:
            cls = widget.winfo_class()
            if cls in ("Frame", "LabelFrame"):
                widget.configure(bg=t["BG"])
            elif cls == "Label":
                widget.configure(bg=widget.master.cget("bg") if hasattr(widget.master, 'cget') else t["BG"])
            elif cls == "Text":
                widget.configure(bg=t["SURFACE2"], fg=t["TEXT"])
            elif cls == "Entry":
                widget.configure(bg=t["SURFACE"], fg=t["TEXT"])
        except Exception:
            pass
        for child in widget.winfo_children():
            self._update_widget_colors(child, t)

    # ── fonts ────────────────────────────────────────────────────────────────
    def _build_fonts(self):
        self.font_title = font.Font(family="Segoe UI", size=16, weight="bold")
        self.font_sub   = font.Font(family="Segoe UI", size=9)
        self.font_label = font.Font(family="Segoe UI", size=10)
        self.font_btn   = font.Font(family="Segoe UI", size=9,  weight="bold")
        self.font_mono  = font.Font(family="Consolas",  size=9)
        self.font_stat  = font.Font(family="Segoe UI", size=10, weight="bold")
        self.font_badge = font.Font(family="Segoe UI", size=8)
        self.font_table = font.Font(family="Segoe UI", size=9)

    # ── header ───────────────────────────────────────────────────────────────
    def _build_header(self):
        t   = self._theme
        hdr = tk.Frame(self, bg=t["SURFACE"], pady=10)
        hdr.pack(fill="x")

        tk.Label(hdr, text="🗂  Smart AI File Organizer",
                 font=self.font_title, bg=t["SURFACE"], fg=t["WHITE"]).pack()
        tk.Label(hdr,
                 text="PDF · DOCX · TXT · XLSX · PPTX · CSV · EML · MSG · ZIP · Images",
                 font=self.font_sub, bg=t["SURFACE"], fg=t["MUTED"]).pack(pady=(2, 0))

    # ── folder row ───────────────────────────────────────────────────────────
    def _build_folder_row(self):
        t     = self._theme
        frame = tk.Frame(self, bg=t["BG"], padx=20, pady=10)
        frame.pack(fill="x")

        tk.Label(frame, text="Target Folder", font=self.font_label,
                 bg=t["BG"], fg=t["TEXT"]).grid(row=0, column=0, sticky="w", pady=(0, 4))

        ef = tk.Frame(frame, bg=t["SURFACE"], highlightthickness=1,
                      highlightbackground=ACCENT)
        ef.grid(row=1, column=0, sticky="ew", padx=(0, 10))
        frame.columnconfigure(0, weight=1)

        self._folder_entry = tk.Entry(
            ef, textvariable=self._folder,
            font=self.font_label, bg=t["SURFACE"], fg=t["TEXT"],
            insertbackground=t["WHITE"], bd=0, relief="flat",
        )
        self._folder_entry.pack(fill="x", padx=8, pady=6)

        tk.Button(frame, text="Browse…", font=self.font_label,
                  bg=ACCENT, fg="white", activebackground=ACCENT_DARK,
                  activeforeground="white", bd=0, padx=14, pady=6,
                  cursor="hand2", relief="flat",
                  command=self._browse).grid(row=1, column=1)

    # ── options row ──────────────────────────────────────────────────────────
    def _build_options_row(self):
        t     = self._theme
        frame = tk.Frame(self, bg=t["BG"], padx=20, pady=2)
        frame.pack(fill="x")

        tk.Label(frame, text="Mode:", font=self.font_label,
                 bg=t["BG"], fg=t["TEXT"]).pack(side="left", padx=(0, 10))

        rb = dict(bg=t["BG"], fg=t["TEXT"], activebackground=t["BG"],
                  activeforeground=t["WHITE"], selectcolor=t["SURFACE"],
                  font=self.font_label, cursor="hand2", bd=0)

        tk.Radiobutton(frame, text="🔍 Preview (safe)",
                       variable=self._dry_run, value=True,
                       **rb).pack(side="left", padx=(0, 18))
        tk.Radiobutton(frame, text="⚡ Organise Now",
                       variable=self._dry_run, value=False,
                       **rb).pack(side="left", padx=(0, 24))
        tk.Checkbutton(frame, text="📂 Include sub-folders",
                       variable=self._recursive,
                       bg=t["BG"], fg=t["TEXT"], activebackground=t["BG"],
                       activeforeground=t["WHITE"], selectcolor=t["SURFACE"],
                       font=self.font_label, cursor="hand2", bd=0,
                       ).pack(side="left", padx=(0, 16))

        tk.Checkbutton(frame, text="✏️ AI Smart Rename",
                       variable=self._smart_rename,
                       bg=t["BG"], fg="#fbbf24", activebackground=t["BG"],
                       activeforeground="#fbbf24", selectcolor=t["SURFACE"],
                       font=self.font_label, cursor="hand2", bd=0,
                       ).pack(side="left")

    # ── file type badges ─────────────────────────────────────────────────────
    def _build_file_types_row(self):
        t     = self._theme
        frame = tk.Frame(self, bg=t["BG"], padx=20, pady=4)
        frame.pack(fill="x")

        tk.Label(frame, text="Supports:", font=self.font_badge,
                 bg=t["BG"], fg=t["MUTED"]).pack(side="left", padx=(0, 6))

        types = [
            ("PDF","#ef4444"),("DOCX","#3b82f6"),("TXT","#10b981"),
            ("XLSX","#22c55e"),("PPTX","#f97316"),("CSV","#06b6d4"),
            ("EML","#8b5cf6"),("MSG","#ec4899"),("ZIP","#64748b"),
            ("PNG","#a855f7"),("JPG","#d946ef"),
        ]
        for label, color in types:
            b = tk.Frame(frame, bg=color, padx=5, pady=2)
            b.pack(side="left", padx=2)
            tk.Label(b, text=label, font=self.font_badge, bg=color, fg="white").pack()

    # ── action buttons ───────────────────────────────────────────────────────
    def _build_action_buttons(self):
        t     = self._theme
        frame = tk.Frame(self, bg=t["BG"], padx=20, pady=6)
        frame.pack(fill="x")

        btn_cfg = dict(font=self.font_btn, bd=0, pady=8, cursor="hand2", relief="flat", fg="white")

        self._btn_run = tk.Button(frame, text="▶  Run", bg=ACCENT,
                                  activebackground=ACCENT_DARK,
                                  padx=20, command=self._run, **btn_cfg)
        self._btn_run.pack(side="left", padx=(0, 8))

        self._btn_watch = tk.Button(frame, text="👁  Watch", bg=WATCH_CLR,
                                    activebackground=WATCH_DARK,
                                    padx=18, command=self._toggle_watch, **btn_cfg)
        self._btn_watch.pack(side="left", padx=(0, 8))

        self._btn_undo = tk.Button(frame, text="↩️  Undo", bg=UNDO_CLR,
                                   activebackground=UNDO_DARK,
                                   padx=18, command=self._undo, **btn_cfg)
        self._btn_undo.pack(side="left", padx=(0, 8))

        tk.Button(frame, text="🗑  Clear", bg=t["SURFACE2"],
                  activebackground=t["BG"], fg=t["TEXT"],
                  padx=14, command=self._clear_log, **{k:v for k,v in btn_cfg.items() if k != "fg"}
                  ).pack(side="left", padx=(0, 8))

        tk.Button(frame, text="⚙️  Categories", bg="#854d0e",
                  activebackground="#713f12", fg="#ffffff",
                  padx=14, command=self._open_categories, **{k:v for k,v in btn_cfg.items() if k != "fg"}
                  ).pack(side="left", padx=(0, 8))

        self._spinner_lbl = tk.Label(frame, text="", font=self.font_label,
                                     bg=t["BG"], fg=WARNING)
        self._spinner_lbl.pack(side="left", padx=10)

    # ── notebook (Log + Results tabs) ────────────────────────────────────────
    def _build_notebook(self):
        t = self._theme
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TNotebook",        background=t["BG"],  borderwidth=0)
        style.configure("TNotebook.Tab",    background=t["SURFACE"], foreground=t["MUTED"],
                         padding=[12, 6],   font=("Segoe UI", 9))
        style.map("TNotebook.Tab",
                  background=[("selected", ACCENT)],
                  foreground=[("selected", "white")])

        outer = tk.Frame(self, bg=t["BG"], padx=20)
        outer.pack(fill="both", expand=True, pady=(0, 4))

        self._nb = ttk.Notebook(outer)
        self._nb.pack(fill="both", expand=True)

        # Tab 1 — Activity Log
        log_frame = tk.Frame(self._nb, bg=t["SURFACE2"])
        self._nb.add(log_frame, text="📋  Activity Log")

        self._log_text = scrolledtext.ScrolledText(
            log_frame, font=self.font_mono, bg=t["SURFACE2"], fg=t["TEXT"],
            insertbackground=t["WHITE"], bd=0, relief="flat",
            state="disabled", wrap="word", padx=10, pady=8,
        )
        self._log_text.pack(fill="both", expand=True)
        for tag, fg in [("INFO", INFO), ("WARNING", WARNING), ("ERROR", ERROR),
                        ("DEBUG", t["MUTED"]), ("SUCCESS", SUCCESS),
                        ("WATCH", WATCH_CLR), ("UNDO", UNDO_CLR), ("TIME", t["MUTED"])]:
            self._log_text.tag_config(tag, foreground=fg)

        # Tab 2 — Results table
        res_frame = tk.Frame(self._nb, bg=t["BG"])
        self._nb.add(res_frame, text="📊  Results")
        self._build_results_table(res_frame)

    def _build_results_table(self, parent):
        t = self._theme
        cols = ("File", "Category", "Confidence", "Status")
        style = ttk.Style()
        style.configure("Results.Treeview",
                        background=t["SURFACE2"], foreground=t["TEXT"],
                        fieldbackground=t["SURFACE2"], font=("Segoe UI", 9),
                        rowheight=26)
        style.configure("Results.Treeview.Heading",
                        background=t["SURFACE"], foreground=t["TEXT"],
                        font=("Segoe UI", 9, "bold"))
        style.map("Results.Treeview", background=[("selected", ACCENT)])

        self._tree = ttk.Treeview(parent, columns=cols, show="headings",
                                   style="Results.Treeview")

        self._tree.heading("File",       text="File")
        self._tree.heading("Category",   text="Category")
        self._tree.heading("Confidence", text="Confidence")
        self._tree.heading("Status",     text="Status")

        self._tree.column("File",       width=300, anchor="w")
        self._tree.column("Category",   width=120, anchor="center")
        self._tree.column("Confidence", width=110, anchor="center")
        self._tree.column("Status",     width=160, anchor="center")

        # Colour tags for rows
        self._tree.tag_configure("low",    foreground=WARNING)
        self._tree.tag_configure("normal", foreground=SUCCESS)
        self._tree.tag_configure("dup",    foreground=t["MUTED"])

        sb = ttk.Scrollbar(parent, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)

        self._tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # Right-click context menu for manual override
        self._ctx_menu = tk.Menu(self, tearoff=0, bg=t["SURFACE"], fg=t["TEXT"])
        self._tree.bind("<Button-3>", self._show_context_menu)
        self._tree.bind("<ButtonRelease-1>", self._on_tree_click)
        self._tooltip = None

        # Preview panel below table
        preview_frame = tk.Frame(parent, bg=t["SURFACE2"], pady=6)
        preview_frame.pack(side="bottom", fill="x", padx=0)

        tk.Label(preview_frame, text="📄 Content Preview",
                 font=self.font_badge, bg=t["SURFACE2"], fg=t["MUTED"]).pack(anchor="w", padx=10)

        self._preview_lbl = tk.Label(
            preview_frame, text="Click any row to preview file content…",
            font=self.font_mono, bg=t["SURFACE2"], fg=t["MUTED"],
            wraplength=700, justify="left", anchor="w",
        )
        self._preview_lbl.pack(fill="x", padx=10, pady=(2, 6))

        # Bottom bar — hint + export button
        bottom = tk.Frame(parent, bg=t["BG"])
        bottom.pack(side="bottom", fill="x", pady=4)

        tk.Label(bottom, text="Right-click any row to override its category",
                 font=self.font_badge, bg=t["BG"], fg=t["MUTED"]).pack(side="left", padx=10)

        tk.Button(bottom, text="📥 Export Excel", font=self.font_badge,
                  bg=t["SURFACE"], fg=t["TEXT"], activebackground=ACCENT,
                  activeforeground="white", bd=0, padx=10, pady=4,
                  cursor="hand2", relief="flat",
                  command=self._export_csv).pack(side="right", padx=10)

    # ── stats bar ────────────────────────────────────────────────────────────
    def _build_stats_bar(self):
        t   = self._theme
        bar = tk.Frame(self, bg=t["SURFACE"], pady=8)
        bar.pack(fill="x", side="bottom")

        self._stat_vars = {
            "total":  tk.StringVar(value="Files: —"),
            "moved":  tk.StringVar(value="Moved: —"),
            "dupes":  tk.StringVar(value="Duplicates: —"),
            "low":    tk.StringVar(value="Low confidence: —"),
            "errors": tk.StringVar(value="Errors: —"),
            "mode":   tk.StringVar(value=""),
        }
        colours = {
            "total": t["TEXT"], "moved": SUCCESS, "dupes": WARNING,
            "low": "#fbbf24", "errors": ERROR, "mode": WATCH_CLR,
        }
        for key, var in self._stat_vars.items():
            tk.Label(bar, textvariable=var, font=self.font_stat,
                     bg=t["SURFACE"], fg=colours[key], padx=12).pack(side="left")

    # ── logging ──────────────────────────────────────────────────────────────
    def _setup_logging(self):
        self._q_handler = QueueHandler(self._log_queue)
        self._q_handler.setLevel(logging.DEBUG)
        root = logging.getLogger()
        root.setLevel(logging.DEBUG)
        root.handlers.clear()
        root.addHandler(self._q_handler)

    def _poll_log_queue(self):
        try:
            while True:
                record = self._log_queue.get_nowait()
                self._write_log(record)
        except queue.Empty:
            pass
        self.after(80, self._poll_log_queue)

    def _write_log(self, record):
        self._log_text.configure(state="normal")
        formatter = logging.Formatter("%(asctime)s")
        timestamp = formatter.format(record)[:19]
        level     = record.levelname
        message   = record.getMessage()

        self._log_text.insert("end", f"{timestamp}  ", "TIME")
        self._log_text.insert("end", f"[{level:<7}]  ", level)

        if "WATCH MODE" in message or "👁" in message:
            tag = "WATCH"
        elif "Restored" in message or "undo" in message.lower():
            tag = "UNDO"
        elif "Run complete" in message or "Moved '" in message or "✅" in message:
            tag = "SUCCESS"
        else:
            tag = level

        self._log_text.insert("end", message + "\n", tag)
        self._log_text.configure(state="disabled")
        self._log_text.see("end")

    # ── browse ───────────────────────────────────────────────────────────────
    def _browse(self):
        folder = filedialog.askdirectory(title="Select folder to organise")
        if folder:
            self._folder.set(folder)
            files = scan_directory(folder, recursive=self._recursive.get())
            self._append_info(f"📁 Selected: {folder}  →  {len(files)} file(s) found.")

    # ── run ──────────────────────────────────────────────────────────────────
    def _run(self):
        if self._watching:
            messagebox.showwarning("Watch Mode", "Stop Watch Mode first.")
            return
        folder = self._folder.get().strip()
        if not folder or not Path(folder).is_dir():
            messagebox.showwarning("Invalid", "Please select a valid folder.")
            return
        if self._running:
            return

        if not self._dry_run.get():
            if not messagebox.askyesno("Confirm", f"Files in:\n  {folder}\n\nwill be MOVED. Continue?"):
                return

        self._running = True
        self._btn_run.configure(state="disabled", text="⏳ Running…")
        self._btn_undo.configure(state="disabled")
        self._spinner_lbl.configure(text="Processing…")
        self._reset_stats()
        self._clear_results()

        threading.Thread(
            target=self._run_organizer,
            args=(folder, self._dry_run.get(), self._recursive.get()),
            daemon=True,
        ).start()

    def _run_organizer(self, folder, dry_run, recursive):
        try:
            # Load API key from config
            import json
            from pathlib import Path as _Path
            cfg = {}
            try:
                cfg = json.loads((_Path(__file__).parent / "config.json").read_text())
            except Exception:
                pass
            api_key = cfg.get("smart_rename", {}).get("api_key", "")

            self._organizer = FileOrganizer(
                target_dir=folder, dry_run=dry_run,
                recursive=recursive, show_progress=False,
                smart_rename=self._smart_rename.get(),
                api_key=api_key,
            )
            stats = self._organizer.run()
            self.after(0, self._on_run_complete, stats, dry_run)
        except Exception as exc:
            logging.getLogger(__name__).error("Error: %s", exc)
            self.after(0, self._on_run_error, str(exc))

    def _on_run_complete(self, stats, dry_run):
        self._running = False
        self._btn_run.configure(state="normal", text="▶  Run")
        self._btn_undo.configure(state="normal")
        self._spinner_lbl.configure(text="")

        label = "Would move" if dry_run else "Moved"
        self._stat_vars["total"].set(f"Files: {stats['total_files']}")
        self._stat_vars["moved"].set(f"{label}: {stats['moved']}")
        self._stat_vars["dupes"].set(f"Duplicates: {stats['duplicates']}")
        self._stat_vars["low"].set(f"Low conf: {stats['low_confidence']}")
        self._stat_vars["errors"].set(f"Errors: {stats['errors']}")

        # Populate results table
        if self._organizer:
            self._populate_results(self._organizer.results)
            # Switch to results tab
            self._nb.select(1)

        done = "DRY RUN complete" if dry_run else "✅ Done"
        self._append_info(f"\n{done} — {stats['moved']} file(s) {'would be moved' if dry_run else 'moved'}.")

        if stats.get("low_confidence", 0) > 0:
            self._append_info(
                f"⚠️  {stats['low_confidence']} file(s) have LOW confidence — "
                f"check the Results tab and right-click to correct if needed."
            )

        if not dry_run and stats["moved"] > 0:
            messagebox.showinfo("Done!",
                f"✅ Organised {stats['moved']} file(s)!\n\n"
                f"Low confidence : {stats['low_confidence']} (review in Results tab)\n"
                f"Duplicates     : {stats['duplicates']}\n"
                f"Errors         : {stats['errors']}")

    def _on_run_error(self, message):
        self._running = False
        self._btn_run.configure(state="normal", text="▶  Run")
        self._btn_undo.configure(state="normal")
        self._spinner_lbl.configure(text="")
        messagebox.showerror("Error", message)

    # ── results table helpers ────────────────────────────────────────────────
    def _populate_results(self, results):
        self._clear_results()
        for fname, category, conf, is_low, _ in results:
            conf_str = f"{conf:.1f}%"
            if is_low:
                status = "⚠️ Low confidence"
                tag    = "low"
            else:
                status = "✅ Confident"
                tag    = "normal"
            self._tree.insert("", "end",
                              values=(fname, category, conf_str, status),
                              tags=(tag,))

    def _clear_results(self):
        for row in self._tree.get_children():
            self._tree.delete(row)

    # ── right-click override ─────────────────────────────────────────────────
    def _show_context_menu(self, event):
        row = self._tree.identify_row(event.y)
        if not row:
            return
        self._tree.selection_set(row)
        values = self._tree.item(row, "values")
        if not values:
            return

        filename = values[0]
        current  = values[1]

        # Build submenu with all categories
        self._ctx_menu.delete(0, "end")
        self._ctx_menu.add_command(
            label=f'Override category for: "{filename}"',
            state="disabled",
        )
        self._ctx_menu.add_separator()

        if self._organizer:
            for cat in self._organizer.classifier.categories:
                if cat != current:
                    self._ctx_menu.add_command(
                        label=f"→ {cat}",
                        command=lambda f=filename, c=cat: self._apply_override(f, c),
                    )

        self._ctx_menu.tk_popup(event.x_root, event.y_root)

    def _apply_override(self, filename: str, new_category: str):
        if not self._organizer:
            return

        success = self._organizer.apply_override(filename, new_category)
        if success:
            # Update the row in the table
            for row in self._tree.get_children():
                vals = self._tree.item(row, "values")
                if vals and vals[0] == filename:
                    self._tree.item(row, values=(
                        filename, new_category, vals[2], "✏️ Manually overridden"
                    ), tags=("normal",))
                    break
            self._append_info(
                f"✏️ Override applied: '{filename}' → {new_category} "
                f"(model has learned from this correction)"
            )
            messagebox.showinfo("Override Applied",
                                f"'{filename}' moved to {new_category}/\n\n"
                                f"The ML model has learned from this correction.")
        else:
            messagebox.showerror("Override Failed",
                                 f"Could not move '{filename}'.\n"
                                 f"The file may have already been moved or deleted.")

    # ── undo ─────────────────────────────────────────────────────────────────
    def _undo(self):
        folder = self._folder.get().strip()
        if not folder or not Path(folder).is_dir():
            messagebox.showwarning("Invalid", "Please select a valid folder.")
            return
        if self._running or self._watching:
            messagebox.showwarning("Busy", "Wait for current operation to finish.")
            return
        log_path = Path(folder) / "organizer.log"
        if not log_path.exists():
            messagebox.showwarning("No Log", "No organizer.log found. Nothing to undo.")
            return
        if not messagebox.askyesno("Confirm Undo",
                                   "Restore all files to their original locations?"):
            return

        self._btn_undo.configure(state="disabled", text="⏳ Undoing…")
        self._btn_run.configure(state="disabled")
        self._append_undo(f"\n↩️  Undo started — {folder}\n")

        threading.Thread(target=self._run_undo, args=(folder,), daemon=True).start()

    def _run_undo(self, folder):
        try:
            stats = undo_moves(folder, dry_run=False)
            self.after(0, self._on_undo_complete, stats)
        except Exception as exc:
            logging.getLogger(__name__).error("Undo error: %s", exc)
            self.after(0, self._on_run_error, str(exc))

    def _on_undo_complete(self, stats):
        self._btn_undo.configure(state="normal", text="↩️  Undo")
        self._btn_run.configure(state="normal")
        self._append_undo(f"\n↩️  Undo complete — {stats['restored']} file(s) restored.")
        messagebox.showinfo("Undo Complete",
                            f"↩️ Restored: {stats['restored']}\n"
                            f"Skipped : {stats['skipped']}\nErrors: {stats['errors']}")

    # ── watch mode ───────────────────────────────────────────────────────────
    def _toggle_watch(self):
        if self._watching:
            self._stop_watch()
        else:
            self._start_watch()

    def _start_watch(self):
        folder = self._folder.get().strip()
        if not folder or not Path(folder).is_dir():
            messagebox.showwarning("Invalid", "Please select a valid folder.")
            return
        self._watching = True
        self._btn_watch.configure(text="⛔ Stop", bg=ERROR, activebackground="#dc2626")
        self._btn_run.configure(state="disabled")
        self._btn_undo.configure(state="disabled")
        self._stat_vars["mode"].set("👁 WATCH MODE ACTIVE")
        self._append_watch(f"\n👁  Watch Mode — monitoring: {folder}\n")
        threading.Thread(target=self._run_watcher,
                         args=(folder, self._recursive.get()), daemon=True).start()

    def _run_watcher(self, folder, recursive):
        try:
            self._watcher = FolderWatcher(target_dir=folder, delay=2.0, recursive=recursive)
            self._watcher.start()
        except Exception as exc:
            logging.getLogger(__name__).error("Watcher error: %s", exc)
            self.after(0, self._on_run_error, str(exc))

    def _stop_watch(self):
        if self._watcher and self._watcher.observer.is_alive():
            self._watcher.observer.stop()
        self._watching = False
        self._btn_watch.configure(text="👁  Watch", bg=WATCH_CLR, activebackground=WATCH_DARK)
        self._btn_run.configure(state="normal")
        self._btn_undo.configure(state="normal")
        self._stat_vars["mode"].set("")
        self._append_watch("\n⛔  Watch Mode stopped.\n")

    # ── helpers ──────────────────────────────────────────────────────────────
    def _append_info(self, text):
        self._log_text.configure(state="normal")
        self._log_text.insert("end", text + "\n", "INFO")
        self._log_text.configure(state="disabled")
        self._log_text.see("end")

    def _append_watch(self, text):
        self._log_text.configure(state="normal")
        self._log_text.insert("end", text + "\n", "WATCH")
        self._log_text.configure(state="disabled")
        self._log_text.see("end")

    def _append_undo(self, text):
        self._log_text.configure(state="normal")
        self._log_text.insert("end", text + "\n", "UNDO")
        self._log_text.configure(state="disabled")
        self._log_text.see("end")

    def _clear_log(self):
        self._log_text.configure(state="normal")
        self._log_text.delete("1.0", "end")
        self._log_text.configure(state="disabled")

    def _reset_stats(self):
        for k, v in [("total","Files: …"),("moved","Moved: …"),
                     ("dupes","Duplicates: …"),("low","Low conf: …"),("errors","Errors: …")]:
            self._stat_vars[k].set(v)


    def _open_categories(self):
        cfg = CONFIG_PATH
        if not cfg.exists():
            # Fall back to config.example.json if config.json doesn't exist
            cfg = ROOT / "config.example.json"
        if not cfg.exists():
            messagebox.showwarning("No Config",
                "config.json not found.\nCopy config.example.json to config.json first.")
            return
        CategoryManager(self, config_path=str(cfg))

    # ── Feature 1: Keyboard shortcuts ──────────────────────────────────────
    def _bind_shortcuts(self):
        self.bind("<Control-r>", lambda e: self._run())
        self.bind("<Control-R>", lambda e: self._run())
        self.bind("<Control-w>", lambda e: self._toggle_watch())
        self.bind("<Control-W>", lambda e: self._toggle_watch())
        self.bind("<Control-z>", lambda e: self._undo())
        self.bind("<Control-Z>", lambda e: self._undo())
        self.bind("<Control-k>", lambda e: self._clear_log())
        self.bind("<Control-K>", lambda e: self._clear_log())

    # ── Feature 2: Export Excel ──────────────────────────────────────────────
    def _export_csv(self):
        if not self._organizer or not self._organizer.results:
            messagebox.showwarning("No Results", "Run the organiser first to get results.")
            return

        filepath = filedialog.asksaveasfilename(
            title="Export Results as Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile="organizer_results.xlsx",
        )
        if not filepath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side)
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Organizer Results"

            # ── Header style ────────────────────────────────────────────────
            hdr_fill = PatternFill("solid", start_color="2a2a3e")
            hdr_font = Font(bold=True, color="e2e8f0", name="Segoe UI", size=10)
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="3a3a5e")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            headers = ["#", "File Name", "Category", "Confidence (%)", "Status"]
            col_widths = [5, 45, 16, 18, 18]

            for col, (h, w) in enumerate(zip(headers, col_widths), 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font    = hdr_font
                cell.fill    = hdr_fill
                cell.alignment = hdr_align
                cell.border  = border
                ws.column_dimensions[get_column_letter(col)].width = w

            ws.row_dimensions[1].height = 22

            # ── Category colours ─────────────────────────────────────────────
            cat_colors = {
                "Finance":  "dcfce7", "Resume":   "dbeafe",
                "AI":       "f0fdf4", "Research": "fef9c3",
                "Personal": "fae8ff", "Legal":    "fee2e2",
                "Medical":  "cffafe", "Other":    "f1f5f9",
            }

            # ── Data rows ────────────────────────────────────────────────────
            for i, (fname, category, conf, is_low, dest) in                     enumerate(self._organizer.results, 1):

                row = i + 1
                status = "⚠ Low confidence" if is_low else "✅ Confident"
                bg = cat_colors.get(category, "ffffff")
                row_fill = PatternFill("solid", start_color=bg)
                row_font = Font(name="Segoe UI", size=9,
                                color="92400e" if is_low else "14532d")

                values = [i, fname, category, round(conf, 1), status]
                aligns = ["center", "left", "center", "center", "center"]

                for col, (val, aln) in enumerate(zip(values, aligns), 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.fill      = row_fill
                    cell.font      = row_font
                    cell.alignment = Alignment(horizontal=aln, vertical="center")
                    cell.border    = border

                ws.row_dimensions[row].height = 18

            # ── Summary row ──────────────────────────────────────────────────
            total_rows = len(self._organizer.results)
            sum_row = total_rows + 2
            ws.cell(row=sum_row, column=1, value="Total").font = Font(bold=True, name="Segoe UI", size=9)
            ws.cell(row=sum_row, column=2, value=f"=COUNTA(B2:B{total_rows+1})").font = Font(bold=True, name="Segoe UI", size=9)
            ws.cell(row=sum_row, column=4, value=f"=AVERAGE(D2:D{total_rows+1})").number_format = "0.0"
            ws.cell(row=sum_row, column=4).font = Font(bold=True, name="Segoe UI", size=9)

            # ── Freeze header ─────────────────────────────────────────────────
            ws.freeze_panes = "A2"

            wb.save(filepath)
            self._append_info(f"\n📥 Exported {total_rows} results → {filepath}")
            messagebox.showinfo("Exported!", f"✅ Saved as Excel:\n{filepath}")

        except ImportError:
            # Fallback to CSV if openpyxl not available
            import csv
            csv_path = filepath.replace(".xlsx", ".csv")
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["File", "Category", "Confidence (%)", "Status"])
                for fname, category, conf, is_low, dest in self._organizer.results:
                    writer.writerow([fname, category, f"{conf:.1f}",
                                     "Low confidence" if is_low else "Confident"])
            messagebox.showinfo("Exported!", f"Saved as CSV:\n{csv_path}")
        except Exception as exc:
            messagebox.showerror("Export Failed", str(exc))

    # ── Feature 3: Click to preview ─────────────────────────────────────────
    def _on_tree_click(self, event):
        row = self._tree.identify_row(event.y)
        if not row:
            return
        values = self._tree.item(row, "values")
        if not values:
            return

        filename, category, confidence, status = values[0], values[1], values[2], values[3]
        t = self._theme

        # Try to get content preview
        text = ""
        if self._organizer:
            text = self._organizer._file_texts.get(filename, "")

        if text and text.strip():
            preview = " ".join(text.split())[:300]
            if len(text.strip()) > 300:
                preview += "…"
            self._preview_lbl.configure(
                text=f"{preview}",
                fg=t["TEXT"]
            )
        else:
            # Show what we know from results
            self._preview_lbl.configure(
                text=f"File: {filename}  |  Category: {category}  |  Confidence: {confidence}  |  {status}",
                fg=t["MUTED"]
            )

    def _hide_tooltip(self):
        pass

    def _on_close(self):
        if self._watching:
            self._stop_watch()
        self.destroy()


def main():
    SmartOrganizerApp().mainloop()


if __name__ == "__main__":
    main()
